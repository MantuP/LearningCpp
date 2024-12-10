###############################################################################################
#File name	: UtilityForExcelFile.py
#Author		: Prabina Pradhan
#Date		: 08/04/2022
#Purpose	:
###############################################################################################
# Pre defined Modules
# Exception need to handle for line number 82 using getTheSheetNames validata 

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from pathlib import Path
import os.path
import datetime
import warnings
#import xlwings as xw
import pandas as pd
import logging

def getExcelColumnVal(p_wbString, p_columnName):
	"""
	this function will get the list of value of a column from excel file.
	"""
	l_col = p_wbString[p_columnName]
	col_list = [l_col[x].value for x in range(len(l_col))]
	return col_list
	
def getExcelColumnValWithRange(p_wbString, p_columnName,p_startRange,p_endRange):
	"""
	this function will get the list of value of a column from excel file with in a range.
	"""
	l_col = p_wbString[p_columnName]
	col_list = [l_col[x].value for x in range(p_startRange,p_endRange,1)]
	return col_list
	
## This function will get the cell id using the column and search string.	
def getIdUsingSearchValueInColumn(p_ws, p_search_string, p_column="A"):
    for row in range(1, p_ws.max_row + 1):
        coordinate = "{}{}".format(p_column, row)
        if p_ws[coordinate].value == p_search_string:
            return p_column, row
    return p_column, None

## This function will get the cell id using the column index and search string.
def getIdUsingSearchValueInColIndex(p_ws, p_search_string, col_idx=1):
    for row in range(1, p_ws.max_row + 1):
        if p_ws[row][col_idx].value == p_search_string:
            return col_idx, row
    return col_idx, None

## This function will get the cell id using the row index and search string.
def getIdUsingSearchValueInIndex(p_ws, p_search_string, row=1):
    for cell in p_ws[row]:
        if cell.value == p_search_string:
            return cell.column, row
    return None, row

def getExcelSheetColAndRowCnt(p_ws):
	"""
	this function will get the number of column and their number of rows in a dictionary
	"""
	l_excelColRowCnt = []
	for col in range(1, p_ws.max_column + 1):
		col_letter = get_column_letter(col)
		max_col_row = len([cell for cell in p_ws[col_letter] if cell.value])
		#print("Column: {}, Row numbers: {}".format(col_letter, max_col_row))
		l_excelColRowCnt.append({col_letter,max_col_row})
	return l_excelColRowCnt

# Split the file name using '.', Only retunr file name not . and file Extension.
def getFileNameWithOutExtension(p_FileName):
	return os.path.basename(p_FileName).split('.')[0]


# Read file from path and name and return its index.
def readExcelFile(p_downloadedPath,p_strExcelFileName):
	l_Xlfile_name = load_workbook(p_downloadedPath+p_strExcelFileName) #added read_only=True, data_only=True
	ls_file_sheet = l_Xlfile_name.active
	return ls_file_sheet

# Read file with sheet name from path and name and return its index.
def readExcelFileUsingShtName(p_downloadedPath,p_strExcelFileName,p_getOrangeSheetName):
	#l_Xlfile_name = load_workbook(p_downloadedPath+p_strExcelFileName,read_only=True, data_only=True) 
	# data_only #helpful to copy the value not formula.
	l_Xlfile_name = load_workbook(p_downloadedPath+p_strExcelFileName,data_only=True)
	ls_file_sheet = l_Xlfile_name.active
	l_SpecifySheet = l_Xlfile_name[p_getOrangeSheetName]
	##raise KeyError("Worksheet {0} does not exist.".format(key))
	if str(l_SpecifySheet) != None and p_getOrangeSheetName in str(l_SpecifySheet):
		return l_SpecifySheet
	else:
		return 0
		
def readExcelFileUsingShtIndex(p_downloadedPath,p_strExcelFileName,p_setSheetIndex=0):
	#l_Xlfile_name = load_workbook(p_downloadedPath+p_strExcelFileName,read_only=True, data_only=True) 
	# data_only #helpful to copy the value not formula.
	l_Xlfile_name = load_workbook(p_downloadedPath+p_strExcelFileName,data_only=True)
	l_Xlfile_name.active = int(p_setSheetIndex)
	if l_Xlfile_name.active != None:
		return l_Xlfile_name.active
	else:
		return 0
#		
def getMaxRowFromSheet(p_SpecifySheet):
	return p_SpecifySheet.max_row

#	
def getMinRowFromSheet(p_SpecifySheet):
	return p_SpecifySheet.max_column

#
def getMaxColFromSheet(p_SpecifySheet):
	return p_SpecifySheet.min_row

#
def getMinColFromSheet(p_SpecifySheet):
	return p_SpecifySheet.min_column

#
def getTheSheetCount(p_downloadedPath,p_strExcelFileName):
	"""
	this function will return sheets count.
	"""
	l_Xlfile_name = load_workbook(p_downloadedPath+p_strExcelFileName)
	ls_file_sheet = l_Xlfile_name.active
	l_NoOfSheetsCnt = len(l_Xlfile_name.sheetnames)
	return l_NoOfSheetsCnt;
	

def getTheSheetNames(p_downloadedPath,p_strExcelFileName):
	"""
	this function will return sheets names.
	"""
	l_Xlfile_name = load_workbook(p_downloadedPath+p_strExcelFileName)
	ls_file_sheet = l_Xlfile_name.active
	##l_NoOfSheetsCnt = len(l_Xlfile_name.sheetnames)
	l_NoOfSheetsName = l_Xlfile_name.sheetnames
	return l_NoOfSheetsName;

# Get the cell value using sheet name and cell name.
def getCellValueUsingSheetNCellId(p_SpecifySheet,p_cellId):
	return p_SpecifySheet[p_cellId].value

def getConvertNumberToDateInDDMMYYYYHH24MISS(p_date_str):
	#return datetime.strptime(p_date_str, '%d/%m/%Y %H:%M:%S').date()  ##from datetime import datetime
	return datetime.datetime.strptime(p_date_str, "%d/%m/%Y %H:%M:%S")

def getConvertStringToList(p_string,p_string_split):
	return list(p_string.split(p_string_split))
	
def createNewExcelFile(p_fullFileName):
	"""
	this function will create a file in respective directory.
	"""
	wb1 = openpyxl.Workbook()
	wb1.save(p_fullFileName)
	return wb1

def writeHeaderIntoNewExcelFile(p_fullFileName,p_OutPutColumnsHeaderList):
	workbook_obj = openpyxl.load_workbook(p_fullFileName)
	sheet_obj = workbook_obj.active
	sheet_obj.append(p_OutPutColumnsHeaderList)
	workbook_obj.save(p_fullFileName)
	
def writeContentIntoNewExcelFile(p_fullFileName,p_DestinationList, p_DialCodesList, p_EffectiveDateList, p_EffectiveEndDateValDate, p_Rate_Zone_TypeSl, p_Traffic_TypeSl, p_Number_TypeSl, p_Zone_IDSl, p_CategorySl, p_Zone_CategorySl, p_Setup_ChargesSl, p_Minimum_Chargeable_UnitsSl, p_PulseSl, p_Unit_QtySl, p_RatePosValList, p_DirectionSl, p_getCurrencyVal, p_InvoiceSl, p_Round_UnitsSl, p_Minimum_Chagreable_AmountSl, p_Tax_PercentSl, p_Free_UnitsSl, p_TierSl, p_IDD_Unit_RateSl, p_RateStatusPosValList,p_getRateChangeIndicatorNeutral):
	li_list = []
	ln_numberOfTarifUploaded = 0
	
	workbook_obj = openpyxl.load_workbook(p_fullFileName)
	sheet_obj = workbook_obj.active
	for i in range(0, len(p_DestinationList), 1):
		if not p_RateStatusPosValList[i] == p_getRateChangeIndicatorNeutral:
			if i < len(p_DestinationList):
				ln_numberOfTarifUploaded = ln_numberOfTarifUploaded + 1
			#if i < 7:
				if len(p_DialCodesList[i]) > 1 and "," in p_DialCodesList[i]:
					#print("Length of diialCode",len(p_DialCodesList[i])," : " ,p_DialCodesList[i] ," : type " ,type(p_DialCodesList[i]))
					#print(ln_numberOfTarifUploaded ,"records are writing ")
					p_DialCodesListElm = p_DialCodesList[i].split(",");
					j = 0
					for j in range(0, len(p_DialCodesListElm), 1):
						if j < len(p_DialCodesListElm):
							li_list.append(p_DestinationList[i])
							li_list.append(p_DialCodesListElm[j])
							li_list.append(p_EffectiveDateList[i])
							li_list.append(p_EffectiveEndDateValDate)
							li_list.append(p_Rate_Zone_TypeSl)
							li_list.append(p_Traffic_TypeSl)
							li_list.append(p_Number_TypeSl)
							li_list.append(p_Zone_IDSl)
							li_list.append(p_CategorySl)
							li_list.append(p_Zone_CategorySl)
							li_list.append(p_Setup_ChargesSl)
							li_list.append(p_Minimum_Chargeable_UnitsSl)
							li_list.append(p_PulseSl)
							li_list.append(p_Unit_QtySl)
							li_list.append(p_RatePosValList[i])
							li_list.append(p_DirectionSl)
							li_list.append(p_getCurrencyVal)
							li_list.append(p_InvoiceSl)
							li_list.append(p_Round_UnitsSl)
							li_list.append(p_Minimum_Chagreable_AmountSl)
							li_list.append(p_Tax_PercentSl)
							li_list.append(p_Free_UnitsSl)
							li_list.append(p_TierSl)
							li_list.append(p_IDD_Unit_RateSl)
							li_list.append(p_RateStatusPosValList[i])
							sheet_obj.append(li_list)
							#print("JJJJ ",li_list)
							li_list.clear() 
							j = j + 1
					#j = 0
					i = i + 1
				else :
					#print(ln_numberOfTarifUploaded ,"records are writing else")
					li_list.append(p_DestinationList[i])
					li_list.append(p_DialCodesList[i])
					li_list.append(p_EffectiveDateList[i])
					li_list.append(p_EffectiveEndDateValDate)
					li_list.append(p_Rate_Zone_TypeSl)
					li_list.append(p_Traffic_TypeSl)
					li_list.append(p_Number_TypeSl)
					li_list.append(p_Zone_IDSl)
					li_list.append(p_CategorySl)
					li_list.append(p_Zone_CategorySl)
					li_list.append(p_Setup_ChargesSl)
					li_list.append(p_Minimum_Chargeable_UnitsSl)
					li_list.append(p_PulseSl)
					li_list.append(p_Unit_QtySl)
					li_list.append(p_RatePosValList[i])
					li_list.append(p_DirectionSl)
					li_list.append(p_getCurrencyVal)
					li_list.append(p_InvoiceSl)
					li_list.append(p_Round_UnitsSl)
					li_list.append(p_Minimum_Chagreable_AmountSl)
					li_list.append(p_Tax_PercentSl)
					li_list.append(p_Free_UnitsSl)
					li_list.append(p_TierSl)
					li_list.append(p_IDD_Unit_RateSl)
					li_list.append(p_RateStatusPosValList[i])
					sheet_obj.append(li_list)
					#print("IIIII",li_list)
					li_list.clear() 
					i = i + 1
			workbook_obj.save(p_fullFileName)
		else:
			pass
			#logging.error("%s file have no changed or new tariff records to update into DB.", p_fullFileName)
			#print("NO",li_list)
			#print(p_RateStatusPosValList[i])
	if ln_numberOfTarifUploaded == 0:
		logging.warning("%s file have no changed or new tariff records to update into DB.", p_fullFileName)
		return 0
		
	return 0
	
def writeContentIntoNewExcelFile_Semecoln(p_fullFileName,p_DestinationList, p_DialCodesList, p_EffectiveDateList, p_EffectiveEndDateValDate, p_Rate_Zone_TypeSl, p_Traffic_TypeSl, p_Number_TypeSl, p_Zone_IDSl, p_CategorySl, p_Zone_CategorySl, p_Setup_ChargesSl, p_Minimum_Chargeable_UnitsSl, p_PulseSl, p_Unit_QtySl, p_RatePosValList, p_DirectionSl, p_getCurrencyVal, p_InvoiceSl, p_Round_UnitsSl, p_Minimum_Chagreable_AmountSl, p_Tax_PercentSl, p_Free_UnitsSl, p_TierSl, p_IDD_Unit_RateSl, p_RateStatusPosValList,p_getRateChangeIndicatorNeutral):
	li_list = []
	ln_numberOfTarifUploaded = 0
	workbook_obj = openpyxl.load_workbook(p_fullFileName)
	sheet_obj = workbook_obj.active
	#print(len(p_DestinationList))
	#print(len(p_RateStatusPosValList))
	for i in range(0, len(p_DestinationList), 1):
		if not p_RateStatusPosValList[i] == p_getRateChangeIndicatorNeutral:
			if i < len(p_DestinationList):
			#if i < 200 :
				ln_numberOfTarifUploaded = ln_numberOfTarifUploaded + 1
				if len(str(p_DialCodesList[i])) > 1 and ";" in str(p_DialCodesList[i]):
					#print("Length of diialCode",len(p_DialCodesList[i])," : " ,p_DialCodesList[i] ," : type " ,type(p_DialCodesList[i]))
					p_DialCodesListElm = p_DialCodesList[i].split(";");
					j = 0
					for j in range(0, len(p_DialCodesListElm), 1):
						if j < len(p_DialCodesListElm):
							li_list.append(p_DestinationList[i])
							li_list.append(p_DialCodesListElm[j])
							li_list.append(p_EffectiveDateList[i])
							li_list.append(p_EffectiveEndDateValDate)
							li_list.append(p_Rate_Zone_TypeSl)
							li_list.append(p_Traffic_TypeSl)
							li_list.append(p_Number_TypeSl)
							li_list.append(p_Zone_IDSl)
							li_list.append(p_CategorySl)
							li_list.append(p_Zone_CategorySl)
							li_list.append(p_Setup_ChargesSl)
							li_list.append(p_Minimum_Chargeable_UnitsSl)
							li_list.append(p_PulseSl)
							li_list.append(p_Unit_QtySl)
							li_list.append(p_RatePosValList[i])
							li_list.append(p_DirectionSl)
							li_list.append(p_getCurrencyVal)
							li_list.append(p_InvoiceSl)
							li_list.append(p_Round_UnitsSl)
							li_list.append(p_Minimum_Chagreable_AmountSl)
							li_list.append(p_Tax_PercentSl)
							li_list.append(p_Free_UnitsSl)
							li_list.append(p_TierSl)
							li_list.append(p_IDD_Unit_RateSl)
							li_list.append(p_RateStatusPosValList[i])
							sheet_obj.append(li_list)
							#print("JJJJ ",li_list)
							li_list.clear() 
						j = j + 1
					#j = 0
					i = i + 1
					#workbook_obj.save(p_fullFileName)
				else :
					li_list.append(p_DestinationList[i])
					li_list.append(p_DialCodesList[i])
					li_list.append(p_EffectiveDateList[i])
					li_list.append(p_EffectiveEndDateValDate)
					li_list.append(p_Rate_Zone_TypeSl)
					li_list.append(p_Traffic_TypeSl)
					li_list.append(p_Number_TypeSl)
					li_list.append(p_Zone_IDSl)
					li_list.append(p_CategorySl)
					li_list.append(p_Zone_CategorySl)
					li_list.append(p_Setup_ChargesSl)
					li_list.append(p_Minimum_Chargeable_UnitsSl)
					li_list.append(p_PulseSl)
					li_list.append(p_Unit_QtySl)
					li_list.append(p_RatePosValList[i])
					li_list.append(p_DirectionSl)
					li_list.append(p_getCurrencyVal)
					li_list.append(p_InvoiceSl)
					li_list.append(p_Round_UnitsSl)
					li_list.append(p_Minimum_Chagreable_AmountSl)
					li_list.append(p_Tax_PercentSl)
					li_list.append(p_Free_UnitsSl)
					li_list.append(p_TierSl)
					li_list.append(p_IDD_Unit_RateSl)
					li_list.append(p_RateStatusPosValList[i])
					sheet_obj.append(li_list)
					#print("IIIII",li_list)
					li_list.clear() 
				i = i + 1
				workbook_obj.save(p_fullFileName)
		else:
			pass
			#logging.error("%s file have no changed or new tariff records to update into DB.", p_fullFileName)
			#print("NO",li_list)
			#print(p_RateStatusPosValList[i])
	if ln_numberOfTarifUploaded == 0:
		logging.warning("%s file have no changed or new tariff records to update into DB.", p_fullFileName)
		return 0
	
	return 0

	
def writeContentIntoNewExcelFile_NoIndicator(p_fullFileName,p_DestinationList, p_DialCodesList, p_EffectiveDateList, p_EffectiveEndDateValDate, p_Rate_Zone_TypeSl, p_Traffic_TypeSl, p_Number_TypeSl, p_Zone_IDSl, p_CategorySl, p_Zone_CategorySl, p_Setup_ChargesSl, p_Minimum_Chargeable_UnitsSl, p_PulseSl, p_Unit_QtySl, p_RatePosValList, p_DirectionSl, p_getCurrencyVal, p_InvoiceSl, p_Round_UnitsSl, p_Minimum_Chagreable_AmountSl, p_Tax_PercentSl, p_Free_UnitsSl, p_TierSl, p_IDD_Unit_RateSl):

	li_list = []
	ln_numberOfTarifUploaded = 0
	ls_Changed = 'Changed'
	workbook_obj = openpyxl.load_workbook(p_fullFileName)
	sheet_obj = workbook_obj.active
	
	for i in range(0, len(p_DestinationList), 1):
		#if i < len(p_DestinationList):
		if i < 200 :
			ln_numberOfTarifUploaded = ln_numberOfTarifUploaded + 1
			if len(str(p_DialCodesList[i])) > 1 and "," in str(p_DialCodesList[i]):
				p_DialCodesListElm = p_DialCodesList[i].split(",")
				j = 0
				for j in range(0, len(p_DialCodesListElm), 1):
					if j < len(p_DialCodesListElm):
						li_list.append(p_DestinationList[i])
						li_list.append(p_DialCodesListElm[j])
						li_list.append(p_EffectiveDateList[i])
						li_list.append(p_EffectiveEndDateValDate)
						li_list.append(p_Rate_Zone_TypeSl)
						li_list.append(p_Traffic_TypeSl)
						li_list.append(p_Number_TypeSl)
						li_list.append(p_Zone_IDSl)
						li_list.append(p_CategorySl)
						li_list.append(p_Zone_CategorySl)
						li_list.append(p_Setup_ChargesSl)
						li_list.append(p_Minimum_Chargeable_UnitsSl)
						li_list.append(p_PulseSl)
						li_list.append(p_Unit_QtySl)
						li_list.append(p_RatePosValList[i])
						li_list.append(p_DirectionSl)
						li_list.append(p_getCurrencyVal)
						li_list.append(p_InvoiceSl)
						li_list.append(p_Round_UnitsSl)
						li_list.append(p_Minimum_Chagreable_AmountSl)
						li_list.append(p_Tax_PercentSl)
						li_list.append(p_Free_UnitsSl)
						li_list.append(p_TierSl)
						li_list.append(p_IDD_Unit_RateSl)
						li_list.append(ls_Changed)
						sheet_obj.append(li_list)
						#print("JJJJ ",li_list)
						li_list.clear() 
					j = j + 1
				i = i + 1
			else :
				li_list.append(p_DestinationList[i])
				li_list.append(p_DialCodesList[i])
				li_list.append(p_EffectiveDateList[i])
				li_list.append(p_EffectiveEndDateValDate)
				li_list.append(p_Rate_Zone_TypeSl)
				li_list.append(p_Traffic_TypeSl)
				li_list.append(p_Number_TypeSl)
				li_list.append(p_Zone_IDSl)
				li_list.append(p_CategorySl)
				li_list.append(p_Zone_CategorySl)
				li_list.append(p_Setup_ChargesSl)
				li_list.append(p_Minimum_Chargeable_UnitsSl)
				li_list.append(p_PulseSl)
				li_list.append(p_Unit_QtySl)
				li_list.append(p_RatePosValList[i])
				li_list.append(p_DirectionSl)
				li_list.append(p_getCurrencyVal)
				li_list.append(p_InvoiceSl)
				li_list.append(p_Round_UnitsSl)
				li_list.append(p_Minimum_Chagreable_AmountSl)
				li_list.append(p_Tax_PercentSl)
				li_list.append(p_Free_UnitsSl)
				li_list.append(p_TierSl)
				li_list.append(p_IDD_Unit_RateSl)
				li_list.append(ls_Changed)
				sheet_obj.append(li_list)
				li_list.clear() 
			i = i + 1
			workbook_obj.save(p_fullFileName)
	
	if ln_numberOfTarifUploaded == 0:
		logging.warning("%s file have no changed or new tariff records to update into DB.", p_fullFileName)
		return 0
	
	return 0
	
def writeContentIntoNewExcelFile_Semecoln_200(p_fullFileName,p_DestinationList, p_DialCodesList, p_EffectiveDateList, p_EffectiveEndDateValDate, p_Rate_Zone_TypeSl, p_Traffic_TypeSl, p_Number_TypeSl, p_Zone_IDSl, p_CategorySl, p_Zone_CategorySl, p_Setup_ChargesSl, p_Minimum_Chargeable_UnitsSl, p_PulseSl, p_Unit_QtySl, p_RatePosValList, p_DirectionSl, p_getCurrencyVal, p_InvoiceSl, p_Round_UnitsSl, p_Minimum_Chagreable_AmountSl, p_Tax_PercentSl, p_Free_UnitsSl, p_TierSl, p_IDD_Unit_RateSl, p_RateStatusPosValList,p_getRateChangeIndicatorNeutral):
	li_list = []
	ln_numberOfTarifUploaded = 0
	workbook_obj = openpyxl.load_workbook(p_fullFileName)
	sheet_obj = workbook_obj.active
	#print(len(p_DestinationList))
	#print(len(p_RateStatusPosValList))
	ln_startZero = 0
	ln_endTwoHan = 200
	for i in range(ln_startZero, ln_endTwoHan, 1):
		if not p_RateStatusPosValList[i] == p_getRateChangeIndicatorNeutral:
			#if i < len(p_DestinationList):
			if i < 200 :
				ln_numberOfTarifUploaded = ln_numberOfTarifUploaded + 1
				if len(str(p_DialCodesList[i])) > 1 and ";" in str(p_DialCodesList[i]):
					#print("Length of diialCode",len(p_DialCodesList[i])," : " ,p_DialCodesList[i] ," : type " ,type(p_DialCodesList[i]))
					p_DialCodesListElm = p_DialCodesList[i].split(";");
					j = 0
					for j in range(0, len(p_DialCodesListElm), 1):
						if j < len(p_DialCodesListElm):
							li_list.append(p_DestinationList[i])
							li_list.append(p_DialCodesListElm[j])
							li_list.append(p_EffectiveDateList[i])
							li_list.append(p_EffectiveEndDateValDate)
							li_list.append(p_Rate_Zone_TypeSl)
							li_list.append(p_Traffic_TypeSl)
							li_list.append(p_Number_TypeSl)
							li_list.append(p_Zone_IDSl)
							li_list.append(p_CategorySl)
							li_list.append(p_Zone_CategorySl)
							li_list.append(p_Setup_ChargesSl)
							li_list.append(p_Minimum_Chargeable_UnitsSl)
							li_list.append(p_PulseSl)
							li_list.append(p_Unit_QtySl)
							li_list.append(p_RatePosValList[i])
							li_list.append(p_DirectionSl)
							li_list.append(p_getCurrencyVal)
							li_list.append(p_InvoiceSl)
							li_list.append(p_Round_UnitsSl)
							li_list.append(p_Minimum_Chagreable_AmountSl)
							li_list.append(p_Tax_PercentSl)
							li_list.append(p_Free_UnitsSl)
							li_list.append(p_TierSl)
							li_list.append(p_IDD_Unit_RateSl)
							li_list.append(p_RateStatusPosValList[i])
							sheet_obj.append(li_list)
							#print("JJJJ ",li_list)
							li_list.clear() 
						j = j + 1
					#j = 0
					i = i + 1
					#workbook_obj.save(p_fullFileName)
				else :
					li_list.append(p_DestinationList[i])
					li_list.append(p_DialCodesList[i])
					li_list.append(p_EffectiveDateList[i])
					li_list.append(p_EffectiveEndDateValDate)
					li_list.append(p_Rate_Zone_TypeSl)
					li_list.append(p_Traffic_TypeSl)
					li_list.append(p_Number_TypeSl)
					li_list.append(p_Zone_IDSl)
					li_list.append(p_CategorySl)
					li_list.append(p_Zone_CategorySl)
					li_list.append(p_Setup_ChargesSl)
					li_list.append(p_Minimum_Chargeable_UnitsSl)
					li_list.append(p_PulseSl)
					li_list.append(p_Unit_QtySl)
					li_list.append(p_RatePosValList[i])
					li_list.append(p_DirectionSl)
					li_list.append(p_getCurrencyVal)
					li_list.append(p_InvoiceSl)
					li_list.append(p_Round_UnitsSl)
					li_list.append(p_Minimum_Chagreable_AmountSl)
					li_list.append(p_Tax_PercentSl)
					li_list.append(p_Free_UnitsSl)
					li_list.append(p_TierSl)
					li_list.append(p_IDD_Unit_RateSl)
					li_list.append(p_RateStatusPosValList[i])
					sheet_obj.append(li_list)
					#print("IIIII",li_list)
					li_list.clear() 
				i = i + 1
				workbook_obj.save(p_fullFileName)
		else:
			pass
			#logging.error("%s file have no changed or new tariff records to update into DB.", p_fullFileName)
			#print("NO",li_list)
			#print(p_RateStatusPosValList[i])
	ln_startZero = ln_startZero + 200
	ln_endTwoHan = ln_endTwoHan + 200
		
	if ln_numberOfTarifUploaded == 0:
		logging.warning("%s file have no changed or new tariff records to update into DB.", p_fullFileName)
		return 0
	
	return 0
#############################################################################
##import Openpyxl
##file=openpyxl.load_workbook('example.xlsx')
##current_sheet=file.get_sheet_by_name('sheet1')  
##Column_C=current_sheet['C']   
##print ( len(column_C))
##data.close()
##data.closed()

## Replace string using search area
## from openpyxl import load_workbook
## from openpyxl.utils import get_column_letter
## import openpyxl
## 
## workbook = openpyxl.load_workbook('RankingInfo.xlsx')
## sheet = workbook["Sheet1"]
## 
## for i in range(1,10):
## 	if sheet['A'+str(i)].value == None:
## 		sheet['A'+str(i)] = '098'
## 	if i == 2:
## 		print (sheet.max_column)
## workbook.save('RankingInfo.xlsx')
#############################################################################	

